"""Check the design rule end-to-end: a per_kW BOS cost changes sizing; a fixed BOS cost only changes LCoE.

Edits temporary copies of the workbook (original untouched), reloads, and re-solves a full year.
"""
import os
import shutil
import tempfile

from openpyxl import load_workbook

from excel_loader import load_inputs, build_optimizer_input, fetch_to_local
from microgrid_highs import optimize_microgrid
from run_microgrid import build_df, _sizing_from_solution

T = 8760  # full year so annual capex and annual grid opex are comparable


def solve(path, n=T):
    inputs = load_inputs(path)
    ns = build_optimizer_input(inputs, build_df(inputs, n), n)
    sol, _, status = optimize_microgrid(ns)
    solar_kW, batt_kWh = _sizing_from_solution(sol)
    return solar_kW, batt_kWh, sol["meta"]["costs"]["LCoE"], status


def bump_bos(src, dst, item, new_unit_cost):
    """Copy workbook src->dst and set bos row `item` unit_cost to new_unit_cost."""
    shutil.copy(src, dst)
    wb = load_workbook(dst)
    ws = wb["bos"]
    header = [c.value for c in ws[1]]
    item_col, cost_col = header.index("item") + 1, header.index("unit_cost") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=item_col).value == item:
            ws.cell(row=row, column=cost_col).value = new_unit_cost
            break
    else:
        raise ValueError(f"bos item {item!r} not found")
    wb.save(dst)


def main():
    base = fetch_to_local()                      # local editable copy (downloads the Sheet if needed)
    tmp = tempfile.mkdtemp()
    per_kw_xlsx = os.path.join(tmp, "per_kw.xlsx")
    fixed_xlsx = os.path.join(tmp, "fixed.xlsx")

    s0, b0, lcoe0, _ = solve(base)
    bump_bos(base, per_kw_xlsx, "DC cabling", 100000.0)      # huge per_kW -> solar uneconomic
    sA, bA, lcoeA, _ = solve(per_kw_xlsx)
    bump_bos(base, fixed_xlsx, "AC protections", 50000.0)    # large fixed -> LCoE up, sizing same
    sB, bB, lcoeB, _ = solve(fixed_xlsx)
    shutil.rmtree(tmp, ignore_errors=True)

    print(f"base:    solar={s0:8.1f} kW  battery={b0:8.1f} kWh  LCoE=${lcoe0:.4f}")
    print(f"per_kW+: solar={sA:8.1f} kW  battery={bA:8.1f} kWh  LCoE=${lcoeA:.4f}")
    print(f"fixed+:  solar={sB:8.1f} kW  battery={bB:8.1f} kWh  LCoE=${lcoeB:.4f}")

    sizing_moved = abs(sA - s0) > 1.0 or abs(bA - b0) > 1.0
    sizing_same = abs(sB - s0) < 0.5 and abs(bB - b0) < 0.5
    lcoe_up = lcoeB > lcoe0 + 1e-9
    print(f"[{'PASS' if sizing_moved else 'FAIL'}] per_kW change moved sizing")
    print(f"[{'PASS' if sizing_same else 'FAIL'}] fixed change left sizing unchanged")
    print(f"[{'PASS' if lcoe_up else 'FAIL'}] fixed change raised LCoE")

    ok = sizing_moved and sizing_same and lcoe_up
    print("RESULT:", "ALL CHECKS PASSED" if ok else "SOME CHECKS FAILED")
    raise SystemExit(0 if ok else 1)


if __name__ == "__main__":
    main()
