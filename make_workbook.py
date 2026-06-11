"""Generate a structured, colour-coded microgrid inputs workbook.

    python make_workbook.py            # writes microgrid_inputs.xlsx

Sheets:
  setup          component selections + ALL project/diesel/fuel parameters (one place)
  solar_panels   PV module catalogue
  batteries      battery catalogue
  inverters      inverter catalogue
  bos            balance-of-system / soft costs (per_kW | per_kWh | fixed | report_only)
  load           1-day hourly load profile (tiled to the year)

Import into Google Sheets (File -> Import -> Upload) to get the same layout + colours.
The loader (excel_loader.load_inputs) reads `setup` as key/value pairs from columns A/B.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "microgrid_inputs.xlsx"

# ---- palette ----------------------------------------------------------------
TITLE   = PatternFill("solid", fgColor="1F4E5F")
SECTION = PatternFill("solid", fgColor="2E86AB")
HEADER  = PatternFill("solid", fgColor="305496")
INPUT   = PatternFill("solid", fgColor="FFF2CC")   # yellow = editable input
GREY    = PatternFill("solid", fgColor="F2F2F2")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
NOTE = Font(italic=True, color="808080")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _title(ws, text, span):
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    c = ws["A1"]; c.value = text; c.fill = TITLE; c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _section(ws, row, text, span):
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    c = ws.cell(row=row, column=1, value=text); c.fill = SECTION; c.font = WHITE_BOLD
    c.alignment = Alignment(horizontal="left", vertical="center")


def _hrow(ws, row, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h); c.fill = HEADER; c.font = WHITE_BOLD; c.border = BORDER


def build_setup(wb):
    ws = wb.create_sheet("setup")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    _title(ws, "MICROGRID INPUTS  —  SETUP", 3)

    # (label, key, value, note). key="" -> section banner; key=None -> column header.
    rows = [
        ("COMPONENT SELECTION  (leave Choice blank to exclude a component)", "", None, None),
        ("Component", None, "Choice", "Notes"),
        ("Solar panel", "solar_panel", "JinkoTiger", "must match a name in 'solar_panels'; blank = no solar"),
        ("Battery", "battery", "Pylontech_US5000", "match 'batteries'; blank = no battery"),
        ("Inverter", "inverter", "SMA_STP50", "match 'inverters'; required if solar"),
        ("Diesel", "diesel", "", "match a model in 'diesel_generators'; blank = no diesel"),
        ("", "", None, None),
        ("PROJECT", "", None, None),
        ("Parameter", None, "Value", "Notes"),
        ("Project lifetime", "lifetime", 20, "years"),
        ("Grid capacity", "grid_capacity_kW", 10000, "kW connection limit"),
        ("Grid price", "grid_price", 0.15, "$/kWh"),
        ("Max grid fraction", "grid_max_fraction", 0.95, "max share of annual load from grid (<1 binds sizing)"),
        ("Latitude", "lat", -3.314732, "PVGIS site latitude (decimal degrees)"),
        ("Longitude", "lon", 37.326358, "PVGIS site longitude (decimal degrees)"),
        ("Solar max", "solar_max_kW", 100000, "upper bound for solar sizing (kW)"),
        ("Battery max", "battery_max_kWh", 100000, "upper bound for battery sizing (kWh)"),
        ("", "", None, None),
        ("FUEL & TANK  (the diesel model is chosen above from 'diesel_generators')", "", None, None),
        ("Parameter", None, "Value", "Notes"),
        ("Tank min", "tank_min_kWh", 0, "kWh"),
        ("Tank max", "tank_max_kWh", 5000, "kWh"),
        ("Tank DoD", "tank_dod", 0.1, "0-1"),
        ("Diesel price", "fuel_usd_per_L", 1.5, "$/L (converts to $/kWh using fixed diesel density)"),
        ("Fuel energy density", "fuel_kWh_per_kg", 11.9, "kWh/kg (informational; price is $/L)"),
    ]
    # IMPORTANT: the loader reads key from column A and value from column B. So we write the
    # machine key in A (hidden-ish) and keep the human label in C? -> simpler: A=key, B=value, C=note,
    # and we show the human label only in the banner. To stay loader-friendly, A holds the KEY.
    r = 3
    sel_value_cells = {}
    for label, key, value, note in rows:
        if key == "" and value is None:           # section banner
            _section(ws, r, label, 3); r += 1; continue
        if key is None:                            # column header row
            _hrow(ws, r, [label, value, note]); r += 1; continue
        a = ws.cell(row=r, column=1, value=key); a.font = BOLD; a.border = BORDER
        b = ws.cell(row=r, column=2, value=value); b.fill = INPUT; b.border = BORDER
        c = ws.cell(row=r, column=3, value=note); c.font = NOTE; c.border = BORDER
        if key in ("solar_panel", "battery", "inverter", "diesel"):
            sel_value_cells[key] = f"B{r}"
        r += 1

    ws.freeze_panes = "A3"
    return ws, sel_value_cells


def build_catalog(wb, name, headers, data):
    ws = wb.create_sheet(name)
    _hrow(ws, 1, headers)
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = GREY
    for j, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = max(12, len(str(h)) + 3)
    ws.freeze_panes = "A2"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    ws_setup, sel_cells = build_setup(wb)

    build_catalog(wb, "solar_panels",
                  ["name", "rating_W", "cost_per_kW", "derating", "lifetime", "maintenance_per_kW"],
                  [["JinkoTiger", 550, 250, 0.90, 25, 2.0],
                   ["LONGi_540", 540, 270, 0.91, 25, 2.5],
                   ["Trina_600", 600, 240, 0.89, 30, 2.0]])
    build_catalog(wb, "batteries",
                  ["name", "module_kWh", "cost_per_kWh", "efficiency", "dod", "c_rate", "lifetime", "maintenance_per_kWh"],
                  [["Pylontech_US5000", 4.8, 367, 0.95, 0.20, 0.5, 12, 3.0],
                   ["BYD_HVM", 2.76, 320, 0.96, 0.10, 0.5, 15, 2.5],
                   ["Tesla_PW", 13.5, 400, 0.94, 0.15, 0.7, 12, 4.0]])
    build_catalog(wb, "inverters",
                  ["name", "rating_kW", "cost_per_kW", "efficiency", "lifetime"],
                  [["SMA_STP50", 50, 120, 0.980, 15],
                   ["Huawei_100KTL", 100, 90, 0.985, 12],
                   ["Fronius_Eco25", 25, 140, 0.975, 15]])
    build_catalog(wb, "diesel_generators",
                  ["name", "cost_per_kW", "efficiency", "max_kW", "lifetime", "maintenance_per_kW"],
                  [["Cat_C9_300", 450, 0.38, 300, 15, 5],
                   ["Cummins_C150", 500, 0.35, 150, 15, 6],
                   ["Perkins_50", 600, 0.33, 50, 12, 8]])
    build_catalog(wb, "bos",
                  ["item", "basis", "unit_cost", "qty", "applies_to", "notes"],
                  [["DC cabling", "per_kW", 35.0, 1, "solar", "DC cable per installed kW"],
                   ["Mounting structure", "per_kW", 60.0, 1, "solar", "racking per kW"],
                   ["Install labor", "per_kW", 45.0, 1, "solar", "labor per kW"],
                   ["Battery DC wiring", "per_kWh", 8.0, 1, "battery", "per kWh"],
                   ["AC protections", "fixed", 4000.0, 1, "project", "breakers / SPD"],
                   ["Engineering+permits", "fixed", 6000.0, 1, "project", "design & permits"],
                   ["Transport", "fixed", 5.0, 700, "project", "unit_cost = $/km, qty = km"],
                   ["Cable length", "report_only", 0.0, 120, "project", "metres (reporting only)"]])
    build_catalog(wb, "load",
                  ["hour", "kWh"],
                  [[h, v] for h, v in enumerate(
                      [50, 50, 50, 50, 50, 60, 120, 200, 260, 300, 320, 300,
                       260, 300, 340, 360, 320, 260, 200, 150, 120, 90, 70, 50])])

    # selection dropdowns (reference the catalogue name columns)
    for key, sheet in [("solar_panel", "solar_panels"), ("battery", "batteries"),
                       ("inverter", "inverters"), ("diesel", "diesel_generators")]:
        dv = DataValidation(type="list", formula1=f"={sheet}!$A$2:$A$100", allow_blank=True)
        ws_setup.add_data_validation(dv)
        dv.add(ws_setup[sel_cells[key]])

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
